import openpyxl


# 1. 读取excel文档
wb = openpyxl.load_workbook('c:\programs/program/dc3.xlsx')

# # 返回一个workbook对象， 有点类似于文件对象;
# print(wb, type(wb))



# 2. 在工作薄中取得工作表
# print(wb.get_sheet_names())
# 返回一个列表， 存储excel表中所有的sheet工作表;
print(wb.sheetnames)

# 返回一个worksheet对象， 返回当前的活动表;
# print(wb.get_active_sheet())
# print(wb.active)



# 3. 获取工作表中， 单元格的信息
# wb.get_sheet_by_name('Sheet1')
sheet = wb['Others']
print(sheet['A1'])
print(sheet['B1'].value)

cell = sheet['B1']
print(cell.row, cell.column)


print(sheet.cell(row=3, column=2))
print(sheet.cell(row=3, column=2).value)
print(sheet.cell(row=3, column=2, value='www'))


# sheet的属性

print(sheet.max_column)
print(sheet.max_row)
print(sheet.title)
sheet.title = 'example'
print(sheet.title)


for row in sheet.rows:
    for cell in row:
        print(cell.value, end='\t')
    print('\n')

wb.save(filename="c:\programs/program/dc3.xlsx")